from openpyxl import Workbook, load_workbook
from db import DB


class ExcelTool:
    def __init__(self, file_name=None):
        self.file_name = file_name

    def write_excel(self, dict_list):
        wb = Workbook()
        ws = wb.active
        for data in dict_list:
            ws.append(list(data.values()))
        wb.save(self.file_name)


    def write_excel(self, header_list, dict_list):
        wb = Workbook()
        ws = wb.active
        ws.append(header_list)
        for data in dict_list:
            matchlist = []
            for key in header_list :
                if key in data :
                    matchlist.append(data[key])
                else :
                    matchlist.append('')

            #ws.append(list(data.values()))
            ws.append(matchlist)
        wb.save(self.file_name)


    @staticmethod
    def get_excel(file_name, sheet_name):
        key_list = list()
        result_list = list()
        work_book = load_workbook(filename=file_name)
        work_sheet = work_book[sheet_name]
        for i, rows in enumerate(work_sheet.rows):
            if i == 0:
                key_list = [row.value for row in rows]
            else:
                result_dict = dict()
                for j, row in enumerate(rows):
                    result_dict[key_list[j]] = row.value
                result_list.append(result_dict)
        return result_list


def get_set_data(data_list):
    result_set = set()
    for r in data_list:
        if r['date'] is not None:
            if r['date'] is not None:
                if type(r['date']) is str:
                    result_set.add((r['page_id'],
                                   r['page_title'].replace('_', ' ') if r['page_title'] is not None else 'None',
                                   r['date']))
                else:
                    result_set.add((r['page_id'],
                                   r['page_title'].replace('_', ' ') if r['page_title'] is not None else 'None',
                                   r['date'].strftime('%Y-%m-%d')))
            else:
                result_set.add((r['page_id'],
                               r['page_title'].replace('_', ' ') if r['page_title'] is not None else 'None',
                               'None'))
    return result_set


if __name__ == '__main__':
    excel_tool = ExcelTool()
    excel1 = get_set_data(excel_tool.get_excel('.xlsx', 'Refined'))
    excel2 = get_set_data(excel_tool.get_excel('.xlsx', 'Refined'))
    excel3 = get_set_data(excel_tool.get_excel('.xlsx', 'Refined'))
    excel4 = get_set_data(excel_tool.get_excel('.xlsx', 'Refined'))
    excel_set = excel1.union(excel2).union(excel3).union(excel4)
    db_set = get_set_data(DB().select('select page_id, page_title, refined_data as date FROM temp_date_of_birth'))
    intersection_set = excel_set.intersection(db_set)
    dif_db = db_set.difference(excel_set)
    dif_excel = excel_set.difference(db_set)
    wb = Workbook()
    ws = wb.create_sheet('intersection')
    for r in intersection_set:
        ws.append(list(r))
    ws = wb.create_sheet('dif_db')
    for r in dif_db:
        ws.append(list(r))
    ws = wb.create_sheet('dif_excel')
    for r in dif_excel:
        ws.append(list(r))
    ws = wb.create_sheet('original')
    for r in db_set:
        ws.append(list(r))
    wb.save('.xlsx')
